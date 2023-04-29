import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl



driver = webdriver.Chrome()


import openpyxl

def createDatabase(filename, sheetname, headers, rows):
    wb = openpyxl.Workbook()

    sheet = wb.active

    sheet.title = sheetname

    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    wb.save(filename)



def takingData(url, classLookingFor, tagHTML = 'div'):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    data = soup.find(tagHTML, class_=classLookingFor)
    return data.get_text()



def openPage(url, classLookingFor, tagHTML ='div'):
    driver.get(url)
    data = takingData(url, classLookingFor, tagHTML)
    driver.quit()
    return data


